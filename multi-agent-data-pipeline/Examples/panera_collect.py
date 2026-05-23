"""
collect.py — Download partial data from already-completed/timed-out Apify runs
and build the final Excel from whatever was captured.

A1–A8 timed out at 5 min each (free-plan limit) but DID collect reviews.
A9–A11 need to be re-run once the monthly limit resets (or plan is upgraded).

Usage:
    export APIFY_TOKEN=your_token_here
    python collect.py
"""

import os
import asyncio
import pandas as pd
from apify_client import ApifyClientAsync
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

APIFY_TOKEN = os.environ.get("APIFY_TOKEN", "")
EXCEL_INPUT = os.path.join(os.path.dirname(__file__), "../Panera_CA_All_Yelp_URLs_Complete.xlsx")
OUTPUT_EXCEL = os.path.join(os.path.dirname(__file__), "Panera_CA_Yelp_Reviews.xlsx")
OUTPUT_CSV   = os.path.join(os.path.dirname(__file__), "Panera_CA_Yelp_Reviews_merged.csv")

# ── Run IDs from the 8 completed agents (A1–A8) ──────────────────────────────
# These ran for 5 minutes and captured partial reviews.
# A9–A11 must be re-run; add their run IDs here once available.
COMPLETED_RUNS = {
    "A1_LA_core":         "p7UkGJJo6XFXGuMIf",
    "A2_LA_east_SGV":     "LK8JWJPtlO2VJiomF",
    "A3_OC_north":        "77tJ5ce4KQc7Ll8iT",
    "A4_San_Diego":       "3GyAJg559FG1Z4sm1",
    "A5_SD_inland_RV":    "DgkM5Gns8HHqjjiQl",
    "A6_IE_SB":           "IbOb0ecVkMrjgedEK",
    "A7_Sacramento":      "FrJGbIIGPPXzkQnMZ",
    "A8_East_Bay_SJ":     "b66quFkCF9OBIcUrf",
    # Add A9–A11 run IDs here once re-run:
    # "A9_SJ_CV":         "...",
    # "A10_Central_Valley":"...",
    # "A11_misc_CA":      "...",
}

# ── URL metadata ──────────────────────────────────────────────────────────────

def load_url_metadata():
    df = pd.read_excel(EXCEL_INPUT, header=3)
    df.columns = ["drop", "num", "region", "yelp_url", "status"]
    df = df.dropna(subset=["yelp_url"])
    df = df[df["yelp_url"].astype(str).str.startswith("http")].copy()
    df["status"] = df["status"].str.strip().str.title()
    df["region"] = df["region"].ffill()
    return df

# ── Data collection ───────────────────────────────────────────────────────────

async def fetch_run_data(client: ApifyClientAsync, label: str, run_id: str) -> list[dict]:
    try:
        run = await client.run(run_id).get()
        if not run:
            print(f"  {label}: run not found ({run_id})")
            return []

        dataset_id = run.get("defaultDatasetId")
        status     = run.get("status", "?")
        if not dataset_id:
            print(f"  {label}: no dataset ID in run metadata")
            return []

        items = []
        async for item in client.dataset(dataset_id).iterate_items():
            item["_agent_label"] = label
            items.append(item)

        print(f"  {label} [{status}]: {len(items)} reviews  (dataset {dataset_id})")
        return items

    except Exception as e:
        print(f"  {label}: ERROR — {e}")
        return []


async def collect_all_runs() -> list[dict]:
    if not APIFY_TOKEN:
        raise ValueError("Set APIFY_TOKEN environment variable.")

    client = ApifyClientAsync(APIFY_TOKEN)
    tasks = [
        fetch_run_data(client, label, run_id)
        for label, run_id in COMPLETED_RUNS.items()
    ]
    print(f"Fetching data from {len(tasks)} completed run(s)...\n")
    results = await asyncio.gather(*tasks)
    return [item for batch in results for item in batch]

# ── Merge + deduplicate ───────────────────────────────────────────────────────

def merge_and_dedupe(all_reviews: list[dict], meta_df: pd.DataFrame) -> pd.DataFrame:
    df = pd.DataFrame(all_reviews)
    before = len(df)

    if "id" in df.columns:
        df = df.drop_duplicates(subset=["id"])
    else:
        cols = [c for c in ["url", "authorName", "date", "text"] if c in df.columns]
        if cols:
            df = df.drop_duplicates(subset=cols)

    print(f"\nDedup: {before} → {len(df)} reviews ({before - len(df)} removed)\n")

    url_to_meta = {
        row["yelp_url"]: (row["region"], row["status"])
        for _, row in meta_df.iterrows()
    }
    url_col = next((c for c in ["businessUrl", "url"] if c in df.columns), None)
    if url_col:
        df["region"]        = df[url_col].map(lambda u: url_to_meta.get(u, ("Unknown", "Unknown"))[0])
        df["outlet_status"] = df[url_col].map(lambda u: url_to_meta.get(u, ("Unknown", "Unknown"))[1])

    return df.reset_index(drop=True)

# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
OPEN_FILL   = PatternFill("solid", start_color="E8F4FD")
BODY_FONT   = Font(name="Arial", size=9)
THIN        = Side(style="thin", color="BDBDBD")

EXPORT_COLS = {
    "businessName":  "Business Name",
    "outlet_status": "Status",
    "region":        "Region",
    "date":          "Review Date",
    "rating":        "Rating",
    "text":          "Review Text",
    "authorName":    "Reviewer",
    "url":           "Business URL",
    "id":            "Review ID",
    "_agent_label":  "Agent",
}
COL_WIDTHS = {
    "Business Name": 28, "Status": 8, "Region": 18, "Review Date": 13,
    "Rating": 7, "Review Text": 60, "Reviewer": 20,
    "Business URL": 40, "Review ID": 22, "Agent": 18,
}


def write_sheet(wb, name, df):
    ws = wb.create_sheet(title=name[:31])
    present = {k: v for k, v in EXPORT_COLS.items() if k in df.columns}
    headers = list(present.values())

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(present.keys(), 1):
            val = row.get(col, "")
            val = "" if pd.isna(val) else val
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = BODY_FONT
            c.fill = OPEN_FILL
            c.border = Border(bottom=THIN)
            c.alignment = Alignment(wrap_text=(col == "text"), vertical="top")

    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(h, 15)
    ws.freeze_panes = "A2"


def write_summary(wb, df, meta_df):
    ws = wb.create_sheet(title="Run Summary", index=0)
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Panera Bread CA — Yelp Review Scrape"
    ws["B2"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = Font(name="Arial", size=9, italic=True, color="666666")

    lf = Font(name="Arial", bold=True, size=10)
    vf = Font(name="Arial", size=10)
    rows = [
        ("Date Range",      "2016-07-18  →  2018-07-18"),
        ("Max Reviews",     "100 per outlet"),
        ("Agents run",      f"{len(COMPLETED_RUNS)} of 11 (A9–A11 pending)"),
        ("Total Reviews",   len(df)),
        ("Outlets covered", df["businessName"].nunique() if "businessName" in df.columns else "—"),
        ("Dedup key",       "review ID"),
    ]
    for i, (label, val) in enumerate(rows, 5):
        ws.cell(row=i, column=2, value=label).font = lf
        ws.cell(row=i, column=3, value=val).font   = vf

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35


def export_excel(df, meta_df):
    wb = Workbook()
    wb.remove(wb.active)
    write_summary(wb, df, meta_df)

    if "region" in df.columns:
        for region in sorted(df["region"].dropna().unique()):
            rdf = df[df["region"] == region].sort_values(
                "date" if "date" in df.columns else df.columns[0], ascending=False
            )
            write_sheet(wb, str(region), rdf)
    else:
        write_sheet(wb, "All Reviews", df)

    wb.save(OUTPUT_EXCEL)
    print(f"Excel saved → {OUTPUT_EXCEL}")

# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    print("=" * 60)
    print("Panera CA — Collecting from completed Apify runs")
    print("=" * 60 + "\n")

    meta_df     = load_url_metadata()
    all_reviews = await collect_all_runs()

    if not all_reviews:
        print("\nNo reviews collected. Check APIFY_TOKEN and run IDs.")
        return

    df = merge_and_dedupe(all_reviews, meta_df)
    df.to_csv(OUTPUT_CSV, index=False)
    print(f"CSV saved  → {OUTPUT_CSV}")
    export_excel(df, meta_df)

    print(f"\nDone — {len(df)} reviews across {df['region'].nunique() if 'region' in df.columns else '?'} regions")
    if len(COMPLETED_RUNS) < 11:
        missing = 11 - len(COMPLETED_RUNS)
        print(f"\n⚠  {missing} agent(s) still pending (A9–A11).")
        print("   Re-run those after your Apify monthly limit resets,")
        print("   add their run IDs to COMPLETED_RUNS, then run collect.py again.")


if __name__ == "__main__":
    asyncio.run(main())
