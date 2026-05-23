"""
Panera Bread CA — Yelp Review Scraper
  • 3 actors running in parallel per wave  (3 × 1024 MB = 3 GB — well under 8 GB free cap)
  • 6 URLs per actor                       (fast enough to finish within 5-min free timeout)
  • Waves run automatically until all 161 open URLs are done
  • Progress saved to CSV after every wave — safe to re-run if interrupted
  • Builds a formatted Excel at the end (one sheet per region)

Usage:
    export APIFY_TOKEN=<your_token>
    python scraper.py

    # Resume after interruption (skips already-scraped URLs):
    python scraper.py --resume
"""

import os, sys, math, asyncio, argparse, json
from datetime import datetime
from pathlib import Path

import pandas as pd
from apify_client import ApifyClientAsync
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

APIFY_TOKEN     = os.environ.get("APIFY_TOKEN", "")
ACTOR_ID        = "tri_angle/yelp-review-scraper"

DATE_FROM       = "2016-07-18"
DATE_TO         = "2018-07-18"
MAX_REVIEWS     = 100
URLS_PER_ACTOR  = 6          # 6 URLs × ~30 s each ≈ 3 min — fits free 5-min timeout
ACTORS_PER_WAVE = 3          # 3 × 1024 MB = 3 GB concurrent memory
WAVE_PAUSE_S    = 8          # seconds between waves

BASE        = Path(__file__).parent
EXCEL_IN    = BASE / "../Panera_CA_All_Yelp_URLs_Complete.xlsx"
PROGRESS_F  = BASE / "progress.json"          # tracks completed URLs
CSV_OUT     = BASE / "reviews_merged.csv"
EXCEL_OUT   = BASE / "Panera_CA_Yelp_Reviews.xlsx"

# ── Load URLs ─────────────────────────────────────────────────────────────────

def load_urls():
    df = pd.read_excel(EXCEL_IN, header=3)
    df.columns = ["drop", "num", "region", "yelp_url", "status"]
    df = df.dropna(subset=["yelp_url"])
    df = df[df["yelp_url"].astype(str).str.startswith("http")].copy()
    df["status"] = df["status"].str.strip().str.title()
    df["region"] = df["region"].ffill()
    open_df = df[df["status"] == "Open"].reset_index(drop=True)
    print(f"  Open outlets: {len(open_df)}")
    return df, open_df

# ── Progress tracking ─────────────────────────────────────────────────────────

def load_progress() -> dict:
    if PROGRESS_F.exists():
        return json.loads(PROGRESS_F.read_text())
    return {"done_urls": [], "dataset_ids": {}}

def save_progress(prog: dict):
    PROGRESS_F.write_text(json.dumps(prog, indent=2))

# ── Single actor run ──────────────────────────────────────────────────────────

async def run_actor(client: ApifyClientAsync, wave_id: int, actor_id: int,
                    urls: list[str]) -> tuple[list[dict], str | None]:
    label = f"Wave {wave_id} / Actor {actor_id}"
    print(f"    [{label}] starting — {len(urls)} URLs")

    run_input = {
        "startUrls":        [{"url": u} for u in urls],
        "maxReviewsPerUrl": MAX_REVIEWS,
        "dateFrom":         DATE_FROM,
        "dateTo":           DATE_TO,
        "language":         "en",
    }
    try:
        run = await client.actor(ACTOR_ID).call(run_input=run_input)
        dataset_id = run["defaultDatasetId"]
        items = []
        async for item in client.dataset(dataset_id).iterate_items():
            item["_wave"]      = wave_id
            item["_actor"]     = actor_id
            item["_scraped_at"] = datetime.utcnow().isoformat()
            items.append(item)
        print(f"    [{label}] ✓ {len(items)} reviews  (dataset {dataset_id})")
        return items, dataset_id
    except Exception as e:
        print(f"    [{label}] ✗ ERROR: {e}")
        return [], None

# ── Wave runner ───────────────────────────────────────────────────────────────

async def run_all_waves(open_df: pd.DataFrame, resume: bool) -> list[dict]:
    prog = load_progress() if resume else {"done_urls": [], "dataset_ids": {}}

    done_set   = set(prog["done_urls"])
    pending    = [u for u in open_df["yelp_url"].tolist() if u not in done_set]

    if not pending:
        print("  All URLs already scraped. Use --resume or delete progress.json.")
        return []

    total_urls   = len(pending)
    total_actors = math.ceil(total_urls / URLS_PER_ACTOR)
    total_waves  = math.ceil(total_actors / ACTORS_PER_WAVE)

    print(f"\n  Pending : {total_urls} URLs")
    print(f"  Actors  : {total_actors}  ({URLS_PER_ACTOR} URLs each)")
    print(f"  Waves   : {total_waves}   ({ACTORS_PER_WAVE} actors per wave)\n")

    # Split pending URLs into actor-sized chunks
    chunks = [pending[i:i+URLS_PER_ACTOR] for i in range(0, total_urls, URLS_PER_ACTOR)]

    all_reviews: list[dict] = []
    client = ApifyClientAsync(APIFY_TOKEN)

    for wave_idx in range(0, len(chunks), ACTORS_PER_WAVE):
        wave_num    = wave_idx // ACTORS_PER_WAVE + 1
        wave_chunks = chunks[wave_idx:wave_idx + ACTORS_PER_WAVE]

        print(f"  ── Wave {wave_num}/{total_waves} ── {len(wave_chunks)} actors × ≤{URLS_PER_ACTOR} URLs ──")

        tasks = [
            run_actor(client, wave_num, i + 1, chunk)
            for i, chunk in enumerate(wave_chunks)
        ]
        results = await asyncio.gather(*tasks)

        wave_reviews = []
        for (items, dataset_id), chunk in zip(results, wave_chunks):
            wave_reviews.extend(items)
            if dataset_id:
                prog["done_urls"].extend(chunk)
                prog["dataset_ids"][dataset_id] = chunk

        all_reviews.extend(wave_reviews)
        save_progress(prog)

        urls_done = len(prog["done_urls"])
        print(f"  Wave {wave_num} done — {len(wave_reviews)} reviews | "
              f"total scraped: {urls_done}/{len(open_df)} URLs\n")

        if wave_idx + ACTORS_PER_WAVE < len(chunks):
            print(f"  Pausing {WAVE_PAUSE_S}s before next wave...")
            await asyncio.sleep(WAVE_PAUSE_S)

    return all_reviews

# ── Merge + attach metadata ───────────────────────────────────────────────────

def merge_and_enrich(all_reviews: list[dict], all_df: pd.DataFrame) -> pd.DataFrame:
    df = pd.DataFrame(all_reviews)

    # Load any previously saved reviews and merge
    if CSV_OUT.exists():
        old = pd.read_csv(CSV_OUT)
        df  = pd.concat([old, df], ignore_index=True)

    before = len(df)
    if "id" in df.columns:
        df = df.drop_duplicates(subset=["id"])
    else:
        cols = [c for c in ["url", "authorName", "date", "text"] if c in df.columns]
        if cols:
            df = df.drop_duplicates(subset=cols)
    print(f"  Dedup: {before} → {len(df)} reviews ({before - len(df)} removed)")

    # Attach region & status from the master URL list
    meta = {row["yelp_url"]: (row["region"], row["status"])
            for _, row in all_df.iterrows()}
    url_col = next((c for c in ["businessUrl", "url"] if c in df.columns), None)
    if url_col:
        df["region"]        = df[url_col].map(lambda u: meta.get(u, ("Unknown","Unknown"))[0])
        df["outlet_status"] = df[url_col].map(lambda u: meta.get(u, ("Unknown","Unknown"))[1])

    return df.reset_index(drop=True)

# ── Excel export ──────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
OPEN_FILL = PatternFill("solid", start_color="E8F4FD")
CLO_FILL  = PatternFill("solid", start_color="FFF3E0")
ROW_FONT  = Font(name="Arial", size=9)
THIN      = Border(bottom=Side(style="thin", color="BDBDBD"))

COLS = {
    "businessName":  ("Business Name", 30),
    "outlet_status": ("Status",         8),
    "region":        ("Region",        18),
    "date":          ("Review Date",   13),
    "rating":        ("Rating",         7),
    "text":          ("Review Text",   60),
    "authorName":    ("Reviewer",      20),
    "url":           ("Business URL",  40),
    "id":            ("Review ID",     22),
    "_wave":         ("Wave #",         8),
    "_actor":        ("Actor #",        8),
}

def _write_sheet(wb, name, df, fill):
    ws = wb.create_sheet(title=name[:31])
    present = [(k, h, w) for k, (h, w) in COLS.items() if k in df.columns]
    for ci, (_, h, _) in enumerate(present, 1):
        c = ws.cell(1, ci, h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, (col, _, _) in enumerate(present, 1):
            val = row.get(col, "")
            val = "" if pd.isna(val) else val
            c = ws.cell(ri, ci, val); c.font = ROW_FONT; c.fill = fill
            c.border = THIN
            c.alignment = Alignment(wrap_text=(col == "text"), vertical="top")
    for ci, (_, _, w) in enumerate(present, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

def _write_summary(wb, df, open_df):
    ws = wb.create_sheet("Run Summary", 0)
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Panera Bread CA — Yelp Reviews"
    ws["B2"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = Font(name="Arial", size=9, italic=True, color="888888")
    lf = Font(name="Arial", bold=True, size=10)
    vf = Font(name="Arial", size=10)
    rows = [
        ("Date range",    f"{DATE_FROM}  →  {DATE_TO}"),
        ("Max reviews",   f"{MAX_REVIEWS} per outlet"),
        ("URLs per actor", URLS_PER_ACTOR),
        ("Actors / wave",  ACTORS_PER_WAVE),
        ("Total outlets",  len(open_df)),
        ("Total reviews",  len(df)),
        ("Unique outlets", df["businessName"].nunique() if "businessName" in df.columns else "—"),
    ]
    for i, (l, v) in enumerate(rows, 5):
        ws.cell(i, 2, l).font = lf
        ws.cell(i, 3, v).font = vf
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30

def export_excel(df, open_df):
    wb = Workbook(); wb.remove(wb.active)
    _write_summary(wb, df, open_df)
    if "region" in df.columns:
        for region in sorted(df["region"].dropna().unique()):
            rdf = df[df["region"] == region]
            has_closed = "outlet_status" in rdf.columns and (rdf["outlet_status"] == "Closed").any()
            _write_sheet(wb, str(region), rdf, CLO_FILL if has_closed else OPEN_FILL)
    else:
        _write_sheet(wb, "All Reviews", df, OPEN_FILL)
    wb.save(EXCEL_OUT)
    print(f"  Excel saved → {EXCEL_OUT}")

# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true",
                        help="Skip already-scraped URLs and continue from where we left off")
    args = parser.parse_args()

    if not APIFY_TOKEN:
        sys.exit("ERROR: set APIFY_TOKEN environment variable first.")

    print("=" * 60)
    print(" Panera CA Yelp Scraper  —  3 actors × 6 URLs per wave")
    print("=" * 60)

    all_df, open_df = load_urls()

    new_reviews = await run_all_waves(open_df, resume=args.resume)

    if not new_reviews and not CSV_OUT.exists():
        print("\nNo reviews collected.")
        return

    print("\nBuilding final dataset...")
    df = merge_and_enrich(new_reviews, all_df)
    df.to_csv(CSV_OUT, index=False)
    print(f"  CSV saved  → {CSV_OUT}")

    print("\nBuilding Excel...")
    export_excel(df, open_df)

    print(f"\n{'='*60}")
    print(f"  Done!  {len(df):,} reviews | {EXCEL_OUT.name}")
    print(f"{'='*60}")

if __name__ == "__main__":
    asyncio.run(main())
