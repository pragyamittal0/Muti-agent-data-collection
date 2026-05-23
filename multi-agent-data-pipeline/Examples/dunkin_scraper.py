"""
Dunkin' CA — Yelp Review Scraper
Following the 15-agent parallel flow, scaled down to 3 actors per wave (free Apify plan).

  • 3 actors running in parallel per wave  (3 × 1024 MB = 3 GB — under 8 GB free cap)
  • 6 URLs per actor                       (fast enough to finish within 5-min free timeout)
  • Waves run automatically until all 176 URLs are done
  • Progress saved after every wave — safe to re-run if interrupted
  • Builds a formatted Excel at the end

Usage:
    export APIFY_TOKEN=<your_token>
    python scraper.py

    # Resume after interruption (skips already-scraped URLs):
    python scraper.py --resume
"""

import os, sys, math, asyncio, argparse, json
from datetime import datetime
from pathlib import Path

import pandas as pd
from apify_client import ApifyClientAsync
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

APIFY_TOKEN     = os.environ.get("APIFY_TOKEN", "")
ACTOR_ID        = "tri_angle/yelp-review-scraper"

DATE_FROM       = "2019-12-15"
DATE_TO         = "2021-12-15"
MAX_REVIEWS     = 200
URLS_PER_ACTOR  = 6          # 6 URLs × ~30 s each ≈ 3 min — fits free 5-min timeout
ACTORS_PER_WAVE = 3          # 3 × 1024 MB = 3 GB concurrent memory
WAVE_PAUSE_S    = 8          # seconds between waves

BASE        = Path(__file__).parent
EXCEL_IN    = BASE / "../Dunkin_CA_Yelp_URLs.xlsx"
PROGRESS_F  = BASE / "progress.json"
CSV_OUT     = BASE / "reviews_merged.csv"
EXCEL_OUT   = BASE / "Dunkin_CA_Yelp_Reviews.xlsx"

# ── Load URLs ─────────────────────────────────────────────────────────────────

def load_urls():
    """Load all Dunkin' CA Yelp URLs from the 'All Yelp URLs' sheet."""
    df = pd.read_excel(EXCEL_IN, sheet_name="All Yelp URLs")
    # Columns are: #, Yelp URL, City, Status, Yelp Slug
    df = df.rename(columns={"Yelp URL": "yelp_url", "City": "city",
                            "Status": "status", "Yelp Slug": "slug"})
    df = df.dropna(subset=["yelp_url"])
    df = df[df["yelp_url"].astype(str).str.startswith("http")].copy()
    print(f"  Loaded {len(df)} URLs ({len(df[df['status']=='Open'])} open, "
          f"{len(df[df['status']!='Open'])} closed/other)")
    return df

# ── Progress tracking ─────────────────────────────────────────────────────────

def load_progress() -> dict:
    if PROGRESS_F.exists():
        return json.loads(PROGRESS_F.read_text())
    return {"done_urls": [], "dataset_ids": {}}

def save_progress(prog: dict):
    PROGRESS_F.write_text(json.dumps(prog, indent=2))

# ── Single actor run ──────────────────────────────────────────────────────────

async def run_actor(client: ApifyClientAsync, wave_id: int, actor_id: int,
                    urls: list) -> tuple:
    label = f"Wave {wave_id} / Actor {actor_id}"
    print(f"    [{label}] starting — {len(urls)} URLs")

    run_input = {
        "startUrls":        [{"url": u} for u in urls],
        "maxReviewsPerUrl": MAX_REVIEWS,
        "dateFrom":         DATE_FROM,
        "dateTo":           DATE_TO,
        "language":         "en",
    }
    try:
        run = await client.actor(ACTOR_ID).call(run_input=run_input)
        dataset_id = run["defaultDatasetId"]
        items = []
        async for item in client.dataset(dataset_id).iterate_items():
            item["_wave"]      = wave_id
            item["_actor"]     = actor_id
            item["_scraped_at"] = datetime.utcnow().isoformat()
            items.append(item)
        print(f"    [{label}] ✓ {len(items)} reviews  (dataset {dataset_id})")
        return items, dataset_id
    except Exception as e:
        print(f"    [{label}] ✗ ERROR: {e}")
        return [], None

# ── Wave runner ───────────────────────────────────────────────────────────────

async def run_all_waves(all_df: pd.DataFrame, resume: bool) -> list:
    prog = load_progress() if resume else {"done_urls": [], "dataset_ids": {}}

    done_set   = set(prog["done_urls"])
    pending    = [u for u in all_df["yelp_url"].tolist() if u not in done_set]

    if not pending:
        print("  All URLs already scraped. Use --resume or delete progress.json.")
        return []

    total_urls   = len(pending)
    total_actors = math.ceil(total_urls / URLS_PER_ACTOR)
    total_waves  = math.ceil(total_actors / ACTORS_PER_WAVE)

    print(f"\n  Pending : {total_urls} URLs")
    print(f"  Actors  : {total_actors}  ({URLS_PER_ACTOR} URLs each)")
    print(f"  Waves   : {total_waves}   ({ACTORS_PER_WAVE} actors per wave)\n")

    chunks = [pending[i:i+URLS_PER_ACTOR] for i in range(0, total_urls, URLS_PER_ACTOR)]

    all_reviews = []
    client = ApifyClientAsync(APIFY_TOKEN)

    for wave_idx in range(0, len(chunks), ACTORS_PER_WAVE):
        wave_num    = wave_idx // ACTORS_PER_WAVE + 1
        wave_chunks = chunks[wave_idx:wave_idx + ACTORS_PER_WAVE]

        print(f"  ── Wave {wave_num}/{total_waves} ── {len(wave_chunks)} actors × ≤{URLS_PER_ACTOR} URLs ──")

        tasks = [
            run_actor(client, wave_num, i + 1, chunk)
            for i, chunk in enumerate(wave_chunks)
        ]
        results = await asyncio.gather(*tasks)

        wave_reviews = []
        for (items, dataset_id), chunk in zip(results, wave_chunks):
            wave_reviews.extend(items)
            if dataset_id:
                prog["done_urls"].extend(chunk)
                prog["dataset_ids"][dataset_id] = chunk

        all_reviews.extend(wave_reviews)
        save_progress(prog)

        # Append to CSV after each wave so partial data is preserved
        if wave_reviews:
            df_wave = pd.DataFrame(wave_reviews)
            if CSV_OUT.exists():
                df_wave.to_csv(CSV_OUT, mode="a", index=False, header=False)
            else:
                df_wave.to_csv(CSV_OUT, index=False)

        urls_done = len(prog["done_urls"])
        print(f"  Wave {wave_num} done — {len(wave_reviews)} reviews | "
              f"total scraped: {urls_done}/{len(all_df)} URLs\n")

        if wave_idx + ACTORS_PER_WAVE < len(chunks):
            print(f"  Pausing {WAVE_PAUSE_S}s before next wave...")
            await asyncio.sleep(WAVE_PAUSE_S)

    return all_reviews

# ── Merge + attach metadata ───────────────────────────────────────────────────

def merge_and_enrich(all_reviews: list, all_df: pd.DataFrame) -> pd.DataFrame:
    # Read full CSV (includes data from previous waves)
    if CSV_OUT.exists():
        df = pd.read_csv(CSV_OUT)
    else:
        df = pd.DataFrame(all_reviews)

    before = len(df)
    if "id" in df.columns:
        df = df.drop_duplicates(subset=["id"])
    else:
        cols = [c for c in ["url", "businessUrl", "reviewerName", "date", "text"]
                if c in df.columns]
        if cols:
            df = df.drop_duplicates(subset=cols)
    print(f"  Dedup: {before} → {len(df)} reviews ({before - len(df)} removed)")

    # Attach status from URL list
    meta = {row["yelp_url"]: row.get("status", "Unknown")
            for _, row in all_df.iterrows()}
    url_col = next((c for c in ["businessUrl", "url"] if c in df.columns), None)
    if url_col:
        df["outlet_status"] = df[url_col].map(lambda u: meta.get(u, "Unknown"))

    return df.reset_index(drop=True)

# ── Excel export ──────────────────────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
OPEN_FILL = PatternFill("solid", start_color="E2EFDA")
CLO_FILL  = PatternFill("solid", start_color="FCE4D6")
ROW_FONT  = Font(name="Calibri", size=10)
THIN      = Border(bottom=Side(style="thin", color="BDBDBD"))

# Required columns: rating, text, businessName, businessUrl, reviewerName, address
COLS = {
    "businessName":    ("Business Name",    30),
    "businessAddress": ("Business Address", 35),
    "outlet_status":   ("Status",            10),
    "date":            ("Review Date",       13),
    "rating":          ("Rating",             7),
    "text":            ("Review Text",       60),
    "reviewerName":    ("Reviewer Name",    22),
    "businessUrl":     ("Business URL",     45),
    "id":              ("Review ID",        22),
    "_wave":           ("Wave #",             7),
    "_actor":          ("Actor #",            7),
}

def _write_sheet(wb, name, df, fill):
    ws = wb.create_sheet(title=name[:31])
    present = [(k, h, w) for k, (h, w) in COLS.items() if k in df.columns]
    for ci, (_, h, _) in enumerate(present, 1):
        c = ws.cell(1, ci, h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, (col, _, _) in enumerate(present, 1):
            val = row.get(col, "")
            val = "" if pd.isna(val) else val
            c = ws.cell(ri, ci, val); c.font = ROW_FONT; c.fill = fill
            c.border = THIN
            c.alignment = Alignment(wrap_text=(col == "text"), vertical="top")
    for ci, (_, _, w) in enumerate(present, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

def _write_summary(wb, df, all_df):
    ws = wb.create_sheet("Run Summary", 0)
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Dunkin' CA — Yelp Reviews"
    ws["B2"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = Font(name="Calibri", size=9, italic=True, color="888888")
    lf = Font(name="Calibri", bold=True, size=11)
    vf = Font(name="Calibri", size=11)
    rows = [
        ("Date range",       f"{DATE_FROM}  →  {DATE_TO}"),
        ("Max reviews",      f"{MAX_REVIEWS} per outlet"),
        ("URLs per actor",   URLS_PER_ACTOR),
        ("Actors / wave",    ACTORS_PER_WAVE),
        ("Total outlets",    len(all_df)),
        ("Total reviews",    len(df)),
        ("Unique outlets",   df["businessName"].nunique() if "businessName" in df.columns else "—"),
    ]
    for i, (l, v) in enumerate(rows, 5):
        ws.cell(i, 2, l).font = lf
        ws.cell(i, 3, v).font = vf
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35

def export_excel(df, all_df):
    wb = Workbook(); wb.remove(wb.active)
    _write_summary(wb, df, all_df)

    # All reviews
    _write_sheet(wb, "All Reviews", df, OPEN_FILL)

    # Open vs Closed
    if "outlet_status" in df.columns:
        odf = df[df["outlet_status"] == "Open"]
        cdf = df[df["outlet_status"] != "Open"]
        if len(odf) > 0:
            _write_sheet(wb, "Open Outlets", odf, OPEN_FILL)
        if len(cdf) > 0:
            _write_sheet(wb, "Closed Outlets", cdf, CLO_FILL)

    wb.save(EXCEL_OUT)
    print(f"  Excel saved → {EXCEL_OUT}")

# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--resume", action="store_true",
                        help="Skip already-scraped URLs and continue from where we left off")
    args = parser.parse_args()

    if not APIFY_TOKEN:
        sys.exit("ERROR: set APIFY_TOKEN environment variable first.")

    print("=" * 60)
    print(" Dunkin' CA Yelp Scraper  —  3 actors × 6 URLs per wave")
    print(f" Date range: {DATE_FROM} → {DATE_TO}")
    print("=" * 60)

    all_df = load_urls()

    new_reviews = await run_all_waves(all_df, resume=args.resume)

    if not new_reviews and not CSV_OUT.exists():
        print("\nNo reviews collected.")
        return

    print("\nBuilding final dataset...")
    df = merge_and_enrich(new_reviews, all_df)
    df.to_csv(CSV_OUT, index=False)
    print(f"  CSV saved  → {CSV_OUT}")

    print("\nBuilding Excel...")
    export_excel(df, all_df)

    print(f"\n{'='*60}")
    print(f"  Done!  {len(df):,} reviews | {EXCEL_OUT.name}")
    print(f"{'='*60}")

if __name__ == "__main__":
    asyncio.run(main())
