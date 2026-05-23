"""Build final Del Taco Yelp Reviews Excel."""
import json, re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE   = Path("/Users/pragyamittal/Downloads/deltaco_yelp_scraper")
CSV    = BASE / "reviews_merged.csv"
URLS_F = BASE / "yelp_urls.json"
OUT    = BASE / "Del_Taco_CA_Yelp_Reviews.xlsx"

# Load reviews
df = pd.read_csv(CSV)
print(f"Total reviews: {len(df)}")
print(f"Unique businesses: {df['businessUrl'].nunique()}")

# Load URL → address mapping (from discovery file)
url_data = json.loads(URLS_F.read_text())
url_to_addr = {}
url_to_city = {}
for rec in url_data.values():
    u = rec.get("yelp_url")
    if u and rec.get("address"):
        # Use first match (multiple stores can map to same Yelp URL)
        if u not in url_to_addr:
            url_to_addr[u] = f"{rec['address']}, {rec['city']}, CA"
            url_to_city[u] = rec['city']

df["businessAddress_resolved"] = df["businessUrl"].map(url_to_addr).fillna(df["businessAddress"]).fillna("")

# Business ID = Yelp slug
df["businessId"] = df["businessUrl"].fillna("").str.split("/biz/").str[-1]

# Reviews per business
biz = (df.groupby("businessUrl")
         .agg(businessName=("businessName", "first"),
              businessAddress=("businessAddress_resolved", "first"),
              businessId=("businessId", "first"),
              num_reviews=("id", "count"),
              avg_rating=("rating", "mean"),
              earliest=("date", "min"),
              latest=("date", "max"))
         .reset_index()
         .sort_values("num_reviews", ascending=False))
biz["avg_rating"] = biz["avg_rating"].round(2)

# Failed URLs (no data)
all_urls = sorted({r["yelp_url"] for r in url_data.values() if r.get("yelp_url")})
have = set(df["businessUrl"].dropna().unique())
failed = [u for u in all_urls if u not in have]

# === Excel ===
HDR_FILL = PatternFill("solid", start_color="C00000")  # Del Taco red
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN = Border(bottom=Side(style="thin", color="BDBDBD"))
GREEN_F = PatternFill("solid", start_color="E2EFDA")
ORANGE_F = PatternFill("solid", start_color="FCE4D6")

def header_row(ws, headers, widths):
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

def write_rows(ws, rows, fill, wrap_cols=None):
    wrap_cols = wrap_cols or set()
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            c = ws.cell(ri, ci, v)
            c.font = BODY_FONT; c.fill = fill; c.border = THIN
            c.alignment = Alignment(vertical="top", wrap_text=(ci in wrap_cols))

wb = Workbook(); wb.remove(wb.active)

# Sheet 1: Summary
ws0 = wb.create_sheet("Summary")
ws0.sheet_view.showGridLines = False
ws0["B2"] = "Del Taco CA — Yelp Reviews"
ws0["B2"].font = Font(name="Calibri", bold=True, size=14, color="C00000")
sf = Font(name="Calibri", bold=True, size=11)
vf = Font(name="Calibri", size=11)
rows = [
    ("Date range",            "2021-03-08 → 2023-03-08"),
    ("Total Yelp URLs",       len(all_urls)),
    ("URLs with reviews",     len(have)),
    ("URLs with NO data",     len(failed)),
    ("Total reviews",         len(df)),
    ("Unique businesses",     int(df["businessUrl"].nunique())),
    ("Avg rating",            round(df["rating"].mean(), 2)),
]
for i, (l, v) in enumerate(rows, 4):
    ws0.cell(i, 2, l).font = sf
    ws0.cell(i, 3, v).font = vf
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 25
ws0.column_dimensions["C"].width = 35

# Sheet 2: Businesses
ws1 = wb.create_sheet("Businesses")
header_row(ws1, ["#", "Business Name", "Business Address", "Business ID", "# Reviews",
                 "Avg Rating", "Earliest Review", "Latest Review", "Business URL"],
           [5, 20, 45, 28, 10, 10, 22, 22, 50])
brows = []
for i, r in enumerate(biz.itertuples(index=False), 1):
    brows.append([i, r.businessName, r.businessAddress, r.businessId, r.num_reviews,
                  r.avg_rating, r.earliest, r.latest, r.businessUrl])
write_rows(ws1, brows, GREEN_F)

# Sheet 3: All Reviews
ws2 = wb.create_sheet("All Reviews")
header_row(ws2, ["#", "Business Name", "Business Address", "Business ID",
                 "Review Date", "Rating", "Review Text", "Reviewer Name", "Business URL"],
           [5, 20, 40, 28, 12, 7, 60, 22, 45])
rrows = []
for i, r in enumerate(df.itertuples(index=False), 1):
    rrows.append([i,
                  getattr(r, "businessName", ""),
                  r.businessAddress_resolved,
                  r.businessId,
                  getattr(r, "date", ""),
                  getattr(r, "rating", ""),
                  getattr(r, "text", ""),
                  getattr(r, "reviewerName", ""),
                  r.businessUrl])
write_rows(ws2, rrows, GREEN_F, wrap_cols={7})

# Sheet 4: Failed URLs
ws3 = wb.create_sheet("Failed URLs")
header_row(ws3, ["#", "Yelp URL", "Reason"], [5, 60, 40])
frows = []
for i, u in enumerate(failed, 1):
    frows.append([i, u, "Yelp DataDome blocked / no reviews in date range"])
write_rows(ws3, frows, ORANGE_F)

wb.save(OUT)
print(f"\nExcel saved → {OUT}")
print(f"  Businesses: {len(biz)}")
print(f"  Reviews: {len(df)}")
print(f"  Failed URLs: {len(failed)}")
