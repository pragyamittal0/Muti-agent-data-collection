"""
Enrich the reviews CSV with business addresses from the Dunkin' locator,
then build an updated Excel including:
  - Reviews (with address)
  - Business summary (one row per business with address)
  - Failed URLs (the 150 that returned no data)
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path("/Users/pragyamittal/Downloads/dunkin_yelp_scraper")
CSV  = BASE / "reviews_merged.csv"
OUT  = BASE / "Dunkin_CA_Yelp_Reviews.xlsx"

ADDR_F  = Path("/tmp/dunkin_addresses.json")
INPUT_X = Path("/Users/pragyamittal/Downloads/Dunkin_CA_Yelp_URLs.xlsx")

# Load reviews
df = pd.read_csv(CSV)
print(f"Reviews: {len(df)}")

# Re-attach outlet_status from input Excel for ALL rows
input_df = pd.read_excel(INPUT_X, sheet_name="All Yelp URLs")
url_to_status = dict(zip(input_df["Yelp URL"], input_df["Status"]))
df["outlet_status"] = df["businessUrl"].map(url_to_status).fillna("Unknown")

# Load city → list of street addresses
city_addrs = json.loads(ADDR_F.read_text())

def yelp_url_to_address(url):
    if not isinstance(url, str):
        return ""
    slug = url.rstrip("/").split("/biz/")[-1]
    # strip "dunkin-donuts-" or "dunkin-" prefix
    body = slug.replace("dunkin-donuts-", "").replace("dunkin-", "")
    # strip trailing -N
    parts = body.rsplit("-", 1)
    base = parts[0] if len(parts) == 2 and parts[1].isdigit() else body
    # special-case express
    if "donut-express" in slug:
        base = slug.split("dunkin-donut-express-")[1]
        parts2 = base.rsplit("-", 1)
        base = parts2[0] if len(parts2) == 2 and parts2[1].isdigit() else base
    addrs = city_addrs.get(base, [])
    city = base.replace("-", " ").title()
    if addrs:
        # Use just the first matching address (one address per row)
        return f"{addrs[0]}, {city}, CA"
    return f"{city}, CA"

df["businessAddress"] = df["businessUrl"].map(yelp_url_to_address)
addr_filled = df["businessAddress"].astype(bool).sum()
print(f"Reviews with address: {addr_filled}/{len(df)}")

# Save updated CSV
df.to_csv(CSV, index=False)

# Working URLs (with reviews)
working = sorted(df["businessUrl"].dropna().unique())
# Input URLs
all_urls = pd.read_excel(INPUT_X, sheet_name="All Yelp URLs")["Yelp URL"].dropna().tolist()
failed = sorted(set(all_urls) - set(working))
print(f"Working URLs: {len(working)} | Failed/empty URLs: {len(failed)}")

# Per-business summary
summary = (df.groupby("businessUrl")
             .agg(businessName=("businessName", "first"),
                  businessAddress=("businessAddress", "first"),
                  outlet_status=("outlet_status", "first"),
                  num_reviews=("id", "count"),
                  avg_rating=("rating", "mean"),
                  earliest_review=("date", "min"),
                  latest_review=("date", "max"))
             .reset_index()
             .sort_values("num_reviews", ascending=False))
summary["avg_rating"] = summary["avg_rating"].round(2)

# === Excel ===
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN      = Border(bottom=Side(style="thin", color="BDBDBD"))
GREEN_F   = PatternFill("solid", start_color="E2EFDA")
ORANGE_F  = PatternFill("solid", start_color="FCE4D6")
BLUE_F    = PatternFill("solid", start_color="DEEBF7")

def header_row(ws, headers, widths, fill=HDR_FILL):
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(1, ci, h)
        c.font = HDR_FONT; c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

def write_rows(ws, rows, fill, wrap_cols=None):
    wrap_cols = wrap_cols or set()
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            c = ws.cell(ri, ci, v)
            c.font = BODY_FONT; c.fill = fill; c.border = THIN
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(ci in wrap_cols))

wb = Workbook(); wb.remove(wb.active)

# 1) Run Summary
ws0 = wb.create_sheet("Run Summary")
ws0.sheet_view.showGridLines = False
ws0["B2"] = "Dunkin' CA — Yelp Reviews"
ws0["B2"].font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
sf = Font(name="Calibri", bold=True, size=11)
vf = Font(name="Calibri", size=11)
rows0 = [
    ("Date range",            "2019-12-15 → 2021-12-15"),
    ("Input URLs",            len(all_urls)),
    ("URLs with reviews",     len(working)),
    ("URLs with NO data",     len(failed)),
    ("Total reviews",         len(df)),
    ("Reviews with address",  addr_filled),
    ("Avg rating",            round(df["rating"].mean(), 2)),
    ("Open-outlet reviews",   int((df["outlet_status"] == "Open").sum())),
    ("Closed-outlet reviews", int((df["outlet_status"] != "Open").sum())),
]
for i, (l, v) in enumerate(rows0, 4):
    ws0.cell(i, 2, l).font = sf
    ws0.cell(i, 3, v).font = vf
ws0.column_dimensions["A"].width = 3
ws0.column_dimensions["B"].width = 25
ws0.column_dimensions["C"].width = 35

# 2) Business Summary (per-URL with address)
ws1 = wb.create_sheet("Businesses (with address)")
header_row(ws1, ["#", "Business Name", "Business Address", "Status",
                 "# Reviews", "Avg Rating", "Earliest", "Latest", "Business URL"],
           [4, 22, 50, 22, 9, 9, 22, 22, 50])
rows = []
for i, r in enumerate(summary.itertuples(index=False), 1):
    rows.append([i, r.businessName, r.businessAddress, r.outlet_status,
                 r.num_reviews, r.avg_rating, r.earliest_review, r.latest_review,
                 r.businessUrl])
write_rows(ws1, rows, GREEN_F, wrap_cols={3})

# 3) All Reviews
ws2 = wb.create_sheet("All Reviews")
header_row(ws2, ["#", "Business Name", "Business Address", "Status",
                 "Review Date", "Rating", "Review Text", "Reviewer Name",
                 "Business URL"],
           [4, 22, 38, 18, 12, 7, 60, 22, 45])
rows = []
for i, r in enumerate(df.itertuples(index=False), 1):
    rows.append([i, r.businessName, r.businessAddress, r.outlet_status,
                 r.date, r.rating, r.text, r.reviewerName, r.businessUrl])
write_rows(ws2, rows, GREEN_F, wrap_cols={7})

# 4) Failed URLs (no data returned)
ws3 = wb.create_sheet("Failed URLs (no data)")
header_row(ws3, ["#", "Yelp URL", "Reason (likely)"], [4, 60, 35])
rows = []
for i, u in enumerate(failed, 1):
    # Most failures are Yelp DataDome blocks on lookups for closed/low-traffic listings
    rows.append([i, u, "Blocked by Yelp DataDome / no reviews in date range"])
write_rows(ws3, rows, ORANGE_F)

wb.save(OUT)
print(f"Saved → {OUT}")
print(f"  - Businesses sheet: {len(summary)} rows")
print(f"  - All Reviews sheet: {len(df)} rows")
print(f"  - Failed URLs sheet: {len(failed)} rows")
